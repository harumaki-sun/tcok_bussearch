import os
import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook

ROUTES_FILE = "routes.xlsx"
OUTPUT_FILE = "bus_data.xlsx"

# ====================================
# 出力ファイルを開く（なければ新規作成）
# ====================================

if os.path.exists(OUTPUT_FILE):
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "BusData"

    headers = [
        "Date",
        "Weekday",
        "CollectedAt",
        "RouteId",
        "SujiId",
        "BusName",
        "PlateNo",
        "StartTime",
        "EndTime"
    ]

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

# ====================================
# 既存データの(Date, SujiId)を読み込む
# ====================================

known_entries = set()

for row in ws.iter_rows(min_row=2, values_only=True):

    date_value = row[0]
    suji_id = row[4]

    if date_value and suji_id:
        known_entries.add((str(date_value), str(suji_id)))

# ====================================
# routes.xlsxを読み込む
# ====================================

routes_wb = load_workbook(ROUTES_FILE)
routes_ws = routes_wb.active

# ====================================
# 今日の日付情報
# ====================================

now = datetime.now()

date_str = now.strftime("%Y-%m-%d")
weekday_str = now.strftime("%a")   # Mon Tue Wed...
collected_at = now.strftime("%Y-%m-%d %H:%M:%S")

# ====================================
# 追記開始行
# ====================================

output_row = ws.max_row + 1

# ====================================
# RouteIdを順番に処理
# ====================================

for row in routes_ws.iter_rows(min_row=2, values_only=True):

    route_id = row[0]

    if not route_id:
        continue

    print(f"処理中: {route_id}")

    url = (
        f"https://kitakyushu.busyohou.jp/"
        f"api/v1/busstop/bus_maps?rid={route_id}"
    )

    try:

        data = requests.get(url, timeout=30).json()

        sujis = data.get("Sujis", [])

        for suji in sujis:

            suji_id = suji.get("SujiId", "")

            # 日付+SujiIdで重複判定
            if (date_str, suji_id) in known_entries:
                continue

            bus = suji.get("Bus", {})

            bus_name = bus.get("BusName", "")
            plate_no = bus.get("PlateNo", "")

            statuses = suji.get("SujiStatus", [])

            if not statuses:
                continue

            first_arrive = statuses[0].get("ArriveTime", "")
            last_arrive = statuses[-1].get("ArriveTime", "")

            if "T" in first_arrive:
                first_arrive = first_arrive.split("T")[1]

            if "T" in last_arrive:
                last_arrive = last_arrive.split("T")[1]

            ws.cell(output_row, 1, date_str)
            ws.cell(output_row, 2, weekday_str)
            ws.cell(output_row, 3, collected_at)
            ws.cell(output_row, 4, route_id)
            ws.cell(output_row, 5, suji_id)
            ws.cell(output_row, 6, bus_name)
            ws.cell(output_row, 7, plate_no)
            ws.cell(output_row, 8, first_arrive)
            ws.cell(output_row, 9, last_arrive)

            known_entries.add((date_str, suji_id))

            print(
                f"追加: {bus_name} "
                f"{suji_id} "
                f"{first_arrive}→{last_arrive}"
            )

            output_row += 1

    except Exception as e:
        print(f"エラー {route_id}: {e}")

# ====================================
# 保存
# ====================================

wb.save(OUTPUT_FILE)

print()
print("保存完了")
print(f"ファイル: {OUTPUT_FILE}")
print(f"総行数: {ws.max_row - 1}")